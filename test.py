import nextcord
from nextcord.ext import commands
from datetime import datetime, timedelta
import random
import openpyxl
import os
from dotenv import load_dotenv
import asyncio


intents = nextcord.Intents.default()
intents.message_content = True  # 메시지 읽기 활성화

load_dotenv()
TOKEN = os.getenv("DISCORD_BOT_TOKEN")

bot = commands.AutoShardedBot(command_prefix="!", intents=intents, help_command=None)

관리자_id = 1355152648914862162   # 관리자 아이디 넣기

# 봇 준비 완료 메시지 출력
@bot.event
async def on_ready():
    print(f'We have logged in as {bot.user}')
    try:
        synced = await bot.sync_application_commands()  # 슬래시 명령어 동기화
        print(f"Synced {len(synced)} commands.")
    except Exception as e:
        print(f"Failed to sync commands: {e}")


@bot.slash_command(name="타임아웃", description="선택한 유저를 타임아웃합니다.", default_member_permissions=nextcord.Permissions(administrator=True))
async def timeout_user(ctx: nextcord.Interaction,
                       멤버: nextcord.Member=nextcord.SlashOption(description="멤버를 입력하세요."),
                       시간: int=nextcord.SlashOption(description="시간을 입력하세요. (분 단위)")):
    
    await ctx.response.defer()  # 응답 지연

     # ✅ 관리자 또는 서버 소유자인지 확인
    if ctx.user.guild_permissions.administrator or ctx.guild.owner_id == ctx.user.id:
        try:
            duration = timedelta(minutes=시간)  #  타임아웃 시간 설정
            await 멤버.timeout(duration, reason="타임아웃 명령어 사용")
            await ctx.followup.send(f"✅ {멤버.mention}님이 {시간}분간 타임아웃 되었습니다.")
        except Exception as e:
            await ctx.followup.send(f"❌ 타임아웃 중 오류가 발생했습니다: {e}")
    else:
        await ctx.followup.send("❌ 관리자 또는 서버 소유자만 사용할 수 있는 명령어입니다!", ephemeral=True)


@bot.slash_command(name="추방", description="유저를 추방함", default_member_permissions=nextcord.Permissions(administrator=True))
async def kick(ctx: nextcord.Interaction, 
               멤버: nextcord.Member = nextcord.SlashOption(description="추방할 멤버를 골라주세요.", required=True),
               사유: str = nextcord.SlashOption(description="사유를 적어주세요", required=False)):
    await ctx.response.defer()

    if ctx.user.guild_permissions.administrator or ctx.guild.owner_id == ctx.user.id:   # 관리자_아이디에 적힌 유저만 사용 가능
    
        if ctx.user.guild_permissions.kick_members:
            await 멤버.kick(reason=사유) # 추방코드
            await ctx.followup.send(f'✅ 추방성공 \n**사유** : {사유}')
        else:
            # 봇이 멤버를 추방할 권한이 없을 떄
            await ctx.followup.send(f"❌구성원을 추방할 권한이 없습니다.", ephemeral=True)
    else:
        # 관리자가 아닌 사람이 이 명령어를 입력하였을 때
        await ctx.followup.send(f"❌이 명령어를 사용할 권한이 없습니다.", ephemeral=True) 


@bot.slash_command(name="서버차단", description="유저를 영구차단함", default_member_permissions=nextcord.Permissions(administrator=True))
async def ban(ctx: nextcord.Interaction, 
              멤버: nextcord.Member = nextcord.SlashOption(description="서버에서 차단할 멤버를 골라주세요.", required=True),
              사유: str = nextcord.SlashOption(description="사유를 적어주세요", required=False)):
    
    await ctx.response.defer()
    
    if ctx.user.guild_permissions.administrator or ctx.guild.owner_id == ctx.user.id:  # 관리자_아이디에 적힌 유저만 사용 가능
        if ctx.user.guild_permissions.ban_members:
            await 멤버.ban(reason=사유)  # 차단코드
            await ctx.followup.send(f'✅ 차단성공 \n**사유** : {사유}')
        else:
            # 봇이 멤버를 차단할 권한이 없을 떄
            await ctx.followup.send(f"❌구성원을 차단할 수 있는 권한이 없습니다.", ephemeral=True)
    else:
        # 관리자가 아닌 사람이 이 명령어를 입력하였을 때
        await ctx.followup.send(f"❌이 명령어를 사용할 권한이 없습니다.", ephemeral=True)


@bot.slash_command(name="가입",description="가입을 할 수 있습니다.")
async def 가입(ctx: nextcord.Interaction, 닉네임: str=nextcord.SlashOption(description="닉네임은 15글자까지 가능합니다.")):
    # 엑셀 파일 경로
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기 (없으면 새로 생성)
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # 새로운 유저 정보 추가
    user_id = str(ctx.user.id)

    # 이미 가입되어 있는지 확인
    for row in sheet.iter_rows(values_only=True):
        if row[0] == user_id:
            await ctx.send("이미 가입되어 있습니다.")
            return
        
    if len(닉네임) > 10: # 닉네임 제한
        await ctx.send("닉네임은 최대 10글자까지만 가능합니다.")
        return

    # 가입되어 있지 않으면 가입 처리
    row = [user_id, 닉네임]
    sheet.append(row)

    # 엑셀 파일 저장
    workbook.save(excel_file)

    await ctx.send(f'✅ {닉네임}님, 가입이 완료되었습니다!')


@bot.slash_command(name="탈퇴",description="탈퇴을 할 수 있습니다.")
async def 탈퇴(ctx):

    # 엑셀 파일 경로
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 유저의 디스코드 아이디 가져오기
        user_id = str(ctx.user.id)

        # 엑셀 파일에서 해당 유저 정보 찾기
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            if row[0] == user_id:
                # 해당 유저 정보를 삭제하고 저장
                sheet.delete_rows(idx)
                workbook.save(excel_file)
                await ctx.send(f"✅ 탈퇴 처리되었습니다.")
                return
        
        # 만약 해당 유저 정보가 없는 경우
        await ctx.send("가입된 정보가 없습니다.")
        
    except FileNotFoundError:
        await ctx.send("가입된 정보가 없습니다.")


@bot.slash_command(name="up지급",description="1000원을 받을 수 있습니다.")
async def 업(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("❌ 가입을 먼저 해주세요.") # 행에 아이디가 존재하지 않을 때
            return


        current_value = sheet.cell(row=target_row_index, column=3).value
        if current_value is None:
            current_value = 0
        current_value = int(current_value)

        # 새로운 값 계산
        new_value = current_value + 1000  # 기존 값에 5000을 더함

        # 값 업데이트
        sheet.cell(row=target_row_index, column=3).value = new_value

        # 엑셀 파일 저장
        workbook.save(excel_file)



        embed = nextcord.Embed(
            title=f"{ctx.user.name}",
            description=f"{ctx.user.mention}님께 1000원을 지급했습니다!",
            color=nextcord.Color(0xF3F781)
        )
        embed.add_field(name="💰 현재 잔액", value=f"{new_value}원", inline=False)
        await ctx.send(embed=embed, ephemeral=False)


    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")


@bot.slash_command(name="범프지급",description="2000원을 받을 수 있습니다.")
async def 범프(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("❌ 가입을 먼저 해주세요.") # 행에 아이디가 존재하지 않을 때
            return


        current_value = sheet.cell(row=target_row_index, column=3).value
        if current_value is None:
            current_value = 0
        current_value = int(current_value)

        # 새로운 값 계산
        new_value = current_value + 2000  # 기존 값에 5000을 더함

        # 값 업데이트
        sheet.cell(row=target_row_index, column=3).value = new_value

        # 엑셀 파일 저장
        workbook.save(excel_file)



        embed = nextcord.Embed(
            title=f"{ctx.user.name}",
            description=f"{ctx.user.mention}님께 2000원을 지급했습니다!",
            color=nextcord.Color(0xF3F781)
        )
        embed.add_field(name="💰 현재 잔액", value=f"{new_value}원", inline=False)
        await ctx.send(embed=embed, ephemeral=False)


    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")



@bot.slash_command(name="추천지급", description="5000원을 받을 수 있습니다.")
async def 추천(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=1).value  
            if cell_value == user_id:
                target_row_index = row_index
                break  # 찾았으므로 중단

        # 가입이 안 되어 있는 경우
        if target_row_index is None:
            await ctx.send("❌ 가입을 먼저 해주세요.", ephemeral=True)
            return

        # 현재 잔액 가져오기
        current_value = sheet.cell(row=target_row_index, column=3).value
        if current_value is None:
            current_value = 0
        current_value = int(current_value)

        # 5000원 지급
        new_value = current_value + 5000

        # 값 업데이트
        sheet.cell(row=target_row_index, column=3).value = new_value

        # 엑셀 파일 저장
        workbook.save(excel_file)

        # ✅ 결과 메시지 출력
        embed = nextcord.Embed(
            title=f"{ctx.user.name}",
            description=f"{ctx.user.mention}님께 5000원을 지급했습니다!",
            color=nextcord.Color(0xF3F781)
        )
        embed.add_field(name="💰 현재 잔액", value=f"{new_value}원", inline=False)

        await ctx.send(embed=embed, ephemeral=False)

    except FileNotFoundError:
        await ctx.send(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        await ctx.send(f"에러 발생: {e}")


@bot.slash_command(name="잔액", description="잔액을 알려줍니다.")
async def 잔액(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("가입을 해주세요.")
            return

        current_value = sheet.cell(row=target_row_index, column=3).value

        current_value = int(current_value)


        # 엑셀 파일 저장
        workbook.save(excel_file)
        embed = nextcord.Embed(
            title=f'{ctx.user.name}',           # 제목과 설명은 임베드에 1개만 추가가 가능합니다
            description='돈 잔액',
            color=nextcord.Color(0xF3F781)  # 색상 코드
        )
        embed.add_field(name='현재 잔액', value=f'{current_value}', inline=False) # 필드
        
        
        await ctx.send(embed=embed, ephemeral=False)


    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")



@bot.slash_command(name="잔액변경", description="유저의 잔액을 변경할 수 있습니다.", default_member_permissions=nextcord.Permissions(administrator=True))
async def 잔액변경(ctx, 
              유저: nextcord.Member = nextcord.SlashOption(description="유저아이디를 입력하세요"), 
              변경할금액: int = nextcord.SlashOption(description="변경할 금액을 입력하세요.")):

    if ctx.user.guild_permissions.administrator or ctx.guild.owner_id == ctx.user.id:
        excel_file = 'data.xlsx'

        try:
            # 엑셀 파일 열기
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 입력한 사용자의 아이디
            user_id = str(유저.id)

            # 사용자 아이디가 있는 행 인덱스 찾기
            target_row_index = None

            for row_index in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_index, column=1).value  
                if cell_value == user_id:
                    target_row_index = row_index
                    break  # 찾았으므로 중단

            # 사용자 아이디가 없을 경우
            if target_row_index is None:
                await ctx.send("❌ 가입이 되어있지 않거나 존재하지 않는 유저입니다.", ephemeral=True)
                return

            # 현재 잔액 가져오기
            current_value = sheet.cell(row=target_row_index, column=3).value
            if current_value is None:
                current_value = 0
            current_value = int(current_value)

            # 새로운 값 계산
            new_value = current_value + 변경할금액  # 잔액 변경

            # 값 업데이트
            sheet.cell(row=target_row_index, column=3).value = new_value

            # 엑셀 파일 저장
            workbook.save(excel_file)

            # ✅ 변경된 잔액을 임베드로 출력
            embed = nextcord.Embed(
                title=f'{ctx.user.name}님의 요청',
                description=f'{유저.mention}님의 잔액 변경 완료!',
                color=nextcord.Color(0xF3F781)  
            )
            embed.add_field(name='변경한 금액', value=f'{변경할금액}원', inline=False)  
            embed.add_field(name='현재 잔액', value=f'{new_value}원', inline=False)  

            await ctx.send(embed=embed, ephemeral=False)

        except FileNotFoundError:
            await ctx.send("❌ 파일을 찾을 수 없습니다. `data.xlsx` 파일이 있는지 확인하세요!", ephemeral=True)
        except Exception as e:
            await ctx.send(f"❌ 오류 발생: {e}", ephemeral=True)


@bot.slash_command(name="메시지삭제", description="입력한 개수만큼 메시지를 삭제합니다.", default_member_permissions=nextcord.Permissions(administrator=True))
async def delete_messages(
    ctx: nextcord.Interaction,
    개수: int = nextcord.SlashOption(description="삭제할 메시지 개수를 입력하세요.", min_value=1, max_value=100)
):
    await ctx.response.defer()  # 응답 지연 방지

    if not ctx.guild.me.guild_permissions.manage_messages:
        return await ctx.followup.send("❌ 봇에게 '메시지 관리' 권한이 없습니다. 서버 설정을 확인하세요.")

    if ctx.user.guild_permissions.administrator or ctx.guild.owner_id == ctx.user.id:
        try:
            deleted = await ctx.channel.purge(limit=개수)
            await ctx.followup.send(f"✅ 최근 {len(deleted)}개의 메시지를 삭제했습니다.", ephemeral=True)
        except nextcord.Forbidden:
            await ctx.followup.send("❌ 메시지 삭제 권한이 부족합니다.")
        except Exception as e:
            await ctx.followup.send(f"❌ 메시지 삭제 중 오류가 발생했습니다.: {e}")
    else:
        await ctx.followup.send("❌ 관리자만 사용할 수 있는 명령어입니다.", ephemeral=True)


def is_on_cooldown(sheet, row, col, cooldown_minutes):
    last_time_str = sheet.cell(row=row, column=col).value
    now = datetime.now()

    if last_time_str:
        try:
            last_time = datetime.strptime(str(last_time_str), "%Y-%m-%d %H:%M:%S")
            remaining = (last_time + timedelta(minutes=cooldown_minutes)) - now
            if remaining.total_seconds() > 0:
                return True, int(remaining.total_seconds())
        except ValueError:
            pass

    return False, 0


@bot.slash_command(name="게시물올리기", description="디스타그램에 게시물을 올립니다.(쿨타임 : 15초)")
async def 게시물올리기(ctx):
    success = [
        "멋진 오운완 사진", "감성 카페에서 찍은 한 컷", "그냥 외모가 원인",
        "해시태그 전략이 제대로 먹혔다", "스토리 공유 이벤트 덕분에 떡상"
    ]
    fail = [
        "감성글 썼다가 감성팔이로 오해받음", "무심코 한 말이 트리거", "과한 보정",
        "정치 얘기 살짝 해버림", "짜증나는 광고같이 보임"
    ]
    neutral = [
        "이상하게 이 사진은 다들 무시함", "알고리즘이 나를 버림", "업로드 시간 실패",
        "너무 자주 올렸더니 피로감 온 듯", "감성 폭발했는데 나만 느낌"
    ]

    result = random.choice(["good", "bad", "neutral"])
    excel_file = "data.xlsx"
    user_id = str(ctx.user.id)

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        row = None
        for i in range(1, ws.max_row + 1):
            if str(ws.cell(i, 1).value) == user_id:
                row = i
                break
        if not row:
            await ctx.send("❗가입하지 않은 사용자입니다.")
            return

        on_cd, secs_left = is_on_cooldown(ws, row, 8, 0.25)  
        if on_cd:
            mins = secs_left // 60
            secs = secs_left % 60
            await ctx.send(f"⏳ 쿨타임입니다. {mins}분 {secs}초 후에 다시 시도해주세요.", ephemeral=True)
            return


        follower = int(ws.cell(row, 4).value or 0)
        like = int(ws.cell(row, 6).value or 0)
        hate = int(ws.cell(row, 7).value or 0)

        if result == "good":
            origin = random.choice(success)
            follower += 10
            like += 30
            msg = f"📈 알고리즘을 탔습니다!\n(원인: {origin})\n+10 Follower / +30 Like"
        elif result == "bad":
            origin = random.choice(fail)
            follower -= 10
            hate += 30
            msg = f"📉 논란의 여지가 있는 사진이네요...\n(원인: {origin})\n-10 Follower / +30 Hate"
        else:
            origin = random.choice(neutral)
            msg = f"😐 이목을 끌지 못했어요..\n(원인: {origin})\n+0 Follower / +0 Like"

        ws.cell(row, 4).value = follower
        ws.cell(row, 6).value = like
        ws.cell(row, 7).value = hate

        ws.cell(row, 8).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb.save(excel_file)

        await ctx.send(embed=nextcord.Embed(title="📸 게시물 업로드", description=msg, color=0xff76c3))
    
    except Exception as e:
        await ctx.send(f"에러 발생: {e}")


@bot.slash_command(name="내피드", description="자신의 디스타그램 피드를 확인합니다.(쿨타임 : 10초)")
async def 내피드(ctx):
    excel_file = "data.xlsx"
    user_id = str(ctx.user.id)

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        row = None
        for i in range(1, ws.max_row + 1):
            if str(ws.cell(i, 1).value) == user_id:
                row = i
                break

        if row is None:
            await ctx.send("❗가입하지 않은 사용자입니다.")
            return

        on_cd, secs_left = is_on_cooldown(ws, row, 9, 0.17)  
        if on_cd:
            mins = secs_left // 60
            secs = secs_left % 60
            await ctx.send(f"⏳ 쿨타임입니다. {mins}분 {secs}초 후에 다시 시도해주세요.", ephemeral=True)
            return

        name = ws.cell(row, 2).value or "이름 없음"
        follower = int(ws.cell(row, 4).value or 0)
        following = int(ws.cell(row, 5).value or 0)
        like = int(ws.cell(row, 6).value or 0)
        hate = int(ws.cell(row, 7).value or 0)

        if follower >= 10000:
            title = "🎤 연예인"
        elif follower >= 5000:
            title = "🌟 인플루언서"
        elif follower >= 1000:
            title = "🔥 라이징스타"
        elif hate >= 10000:
            title = "💀 혐오유발자"
        elif hate >= 5000:
            title = "🦇 다크나이트"
        elif hate >= 1000:
            title = "💢 불행전달자"
        else:
            title = "👤 일반인"

        embed = nextcord.Embed(title="📱 내 디스타그램 피드", color=0xbf74fd)
        embed.add_field(name="이름", value=name, inline=False)
        embed.add_field(name="📈 팔로워", value=str(follower), inline=True)
        embed.add_field(name="📉 팔로잉", value=str(following), inline=True)
        embed.add_field(name="❤️ 좋아요", value=str(like), inline=True)
        embed.add_field(name="💔 싫어요", value=str(hate), inline=True)
        embed.add_field(name="🏷️ 칭호", value=title, inline=False)

        ws.cell(row, 9).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb.save(excel_file)

        await ctx.send(embed=embed)

    except Exception as e:
        await ctx.send(f"에러 발생: {e}")


@bot.slash_command(name="이벤트", description="랜덤 이벤트가 발생합니다(쿨타임 : 15분)")
async def 이벤트(ctx):
    import random
    excel_file = "data.xlsx"
    user_id = str(ctx.user.id)

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        row = None
        for i in range(1, ws.max_row + 1):
            if str(ws.cell(i, 1).value) == user_id:
                row = i
                break

        if row is None:
            await ctx.send("❗가입하지 않은 사용자입니다.")
            return

        on_cd, secs_left = is_on_cooldown(ws, row, 10, 15)  # 3분 쿨타임 예시
        if on_cd:
            mins = secs_left // 60
            secs = secs_left % 60
            await ctx.send(f"⏳ 쿨타임입니다. {mins}분 {secs}초 후에 다시 시도해주세요.", ephemeral=True)
            return

        follower = int(ws.cell(row, 4).value or 0)
        following = int(ws.cell(row, 5).value or 0)
        like = int(ws.cell(row, 6).value or 0)
        hate = int(ws.cell(row, 7).value or 0)

        events = [
            ("📺 방송에 출연했어요!", 1000, 0, 1000, 0, 0.1),
            ("💸 팔로워 구매에 홀렸어요...", 200, 0, 0, 100, 5),
            ("🔓 해킹을 당했어요 (팔로잉)", -50, 200, 100, 0, 5),
            ("📈 릴스가 떡상했어요!", 100, 0, 500, 0, 10),
            ("❌ 해킹을 당했어요 (계정)", -follower, -following, -like, -hate, 0.1),
            ("🏢 기획사에 들어갔어요!", 500, 0, 500, 0, 0.4),
            ("🧹 팔로잉을 정리했어요!", 0, -100, 0, 50, 0.4),
            ("🗯️ 혐오발언을 했어요...", -200, 0, 0, 500, 4.5),
            ("❤️ 기부 사진을 올렸어요!", 200, 0, 1000, 0, 4.5),
            ("📶 소소한 오름", 1, 0, 1, 0, 45),
            ("💤 아무 일도 없었어요", 0, 0, 0, 0, 45),
        ]

        weights = [e[5] for e in events]
        selected = random.choices(events, weights=weights, k=1)[0]
        name, f_change, fg_change, l_change, h_change, _ = selected
  
        if name == "❌ 해킹을 당했어요 (계정)":
            follower, following, like, hate = 0, 0, 0, 0
        else:
            follower += f_change
            following += fg_change
            like += l_change
            hate += h_change

        # 값 저장
        ws.cell(row, 4, follower)
        ws.cell(row, 5, following)
        ws.cell(row, 6, like)
        ws.cell(row, 7, hate)
        ws.cell(row, 10).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb.save(excel_file)

        embed = nextcord.Embed(title="🎲 이벤트 발생!", description=name, color=0xffdf7c)
        if name == "💤 아무 일도 없었어요":
            embed.add_field(name="🥱", value="정말 아무 일도 없었어요...", inline=False)
        else:
            embed.add_field(name="📊 변화량", value=
                f"📈 팔로워: {f_change:+}\n"
                f"📉 팔로잉: {fg_change:+}\n"
                f"❤️ 좋아요: {l_change:+}\n"
                f"💔 싫어요: {h_change:+}", inline=False
            )

        await ctx.send(embed=embed)

    except Exception as e:
        await ctx.send(f"에러 발생: {e}")


@bot.command(name="닉네임변경")
async def 닉네임변경(ctx, *, 새_닉네임: str):
    # ✅ 봇이 닉네임 변경 권한을 가지고 있는지 확인
    if not ctx.guild.me.guild_permissions.manage_nicknames:
        return await ctx.send("❌ 봇에게 '닉네임 변경' 권한이 없습니다. 서버 설정을 확인하세요.")

    try:
        await ctx.author.edit(nick=새_닉네임)
        await ctx.send(f"✅ {ctx.author.mention}님의 닉네임이 `{새_닉네임}`(으)로 변경되었습니다.")
    except nextcord.Forbidden:
        await ctx.send("❌ 닉네임 변경 권한이 부족합니다.")
    except Exception as e:
        await ctx.send(f"❌ 닉네임 변경 중 오류가 발생했습니다: {e}")


@bot.command(name="치우")  
async def 치우(ctx):
    await ctx.send("서버 내 최고 알파메일") 

@bot.command(name="낭만전사")  
async def 낭만전사(ctx):
    await ctx.send('서버 내 핵폭탄급 존재감')  
    
@bot.command(name="수빈")  
async def 수빈(ctx):
    await ctx.send("서버 내 최강 미녀") 

@bot.command(name="인천나얼")  
async def 인천나얼(ctx):
    await ctx.send("서버 내 핵심인재") 

@bot.command(name="봄")  
async def 봄(ctx):
    await ctx.send('봄이는 수비니를 조아해')  

@bot.command(name="어서오세요") # 명령
async def 어서오세요(ctx):
    ran = random.randint(0,3)  # 랜덤으로 보낼 답변의 갯수
    if ran == 0:         
        r = "반가워요"  
    if ran == 1:  
        r = "환영해요"  
    if ran == 2:
        r = "음챗해요"
    if ran == 3:
        r = "게임해요"
    await ctx.channel.send(r)  # 변수 r의 값을 보냄

@bot.event
async def on_ready():
    print(f'We have logged in as {bot.user}')
    print("등록된 명령어 목록:", [cmd.name for cmd in bot.commands])

bot.run(TOKEN) #토큰
